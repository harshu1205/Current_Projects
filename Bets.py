from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import openpyxl
import time
from openpyxl.styles import Color, PatternFill, Font, Border
import pandas as pd

service = Service()
options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# options.add_argument("--headless")
driver = webdriver.Chrome(service=service, options=options)

Link = "https://www.lineupexperts.com/DFS/Underdog"
driver.get(Link)
content = driver.page_source
soup = BeautifulSoup(content, features="html.parser")

Workbook = openpyxl.Workbook()
Sheet = Workbook.active

##########################LINEUP_EXPERTS_DATA#########################################
Datatable = {}
DataframeTable = {}
DKLinks = {
    "Batter Strikeouts": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=batter-props&subcategory=strikeouts",
    "Earned Runs Allowed": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=pitcher-props&subcategory=earned-runs",
    "H + R + RBIS": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=batter-props&subcategory=hits-+-runs-+-rbis",
    "Hits Allowed": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=pitcher-props&subcategory=hits-allowed",
    "Outs recorded": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=pitcher-props&subcategory=outs-recorded",
    "Runs Scored": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=batter-props&subcategory=runs-scored",
    "Strikeouts": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=pitcher-props&subcategory=strikeouts",
    "Total Bases": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=batter-props&subcategory=total-bases",
    "Walks Allowed": "https://sportsbook.draftkings.com/leagues/baseball/mlb?category=pitcher-props&subcategory=walks",
}

Categories = []
Links = []

MainTable = driver.find_element(By.ID, 'TableOutput')
TableChildren = MainTable.find_elements(By.CLASS_NAME, 'text-left')

for i in range(len(TableChildren)):
    if i % 2 == 0 and i != 0:
        text = TableChildren[i].get_attribute("innerText")
        Categories.append(text)
        Datatable[text] = {}

        LinkElement = TableChildren[i].find_element(By.TAG_NAME, "a")
        Links.append(LinkElement.get_attribute("href"))

for i in range(len(Links)):
    link = Links[i]
    driver.get(link)
    content = driver.page_source
    soup = BeautifulSoup(content, features="html.parser")

    Category = Categories[i]

    TableElement = driver.find_elements(By.TAG_NAME, 'table')[1]
    TableBody = TableElement.find_element(By.TAG_NAME, 'tbody')
    TRs = TableBody.find_elements(By.TAG_NAME, 'tr')

    CurrentProp_Datalist = []
    for j in range(len(TRs)):
        TR = TRs[j]

        if j % 4 == 0:
            C = TR.find_elements("xpath", '*')
            InfoList = [C[1].get_attribute("innerText"), C[4].get_attribute("innerText"), C[5].get_attribute("innerText"), C[6].get_attribute("innerText")]
            CurrentProp_Datalist.append(InfoList)
        elif j % 4 == 3:
            C = TR.find_elements("xpath", '*')
            Datalist = []
            for Col in C:
                Datalist.append(Col.get_attribute("innerText"))
            CurrentProp_Datalist.append(Datalist)

    if len(CurrentProp_Datalist) >= 6:
        CurrentProp_Datalist = CurrentProp_Datalist[:6]

    Datatable[Category] = CurrentProp_Datalist

for Prop in Datatable:
    DataframeTable[Prop] = []
    DFTable = []
    for List in Datatable[Prop]:
        DFTable.append(List)

    DF = pd.DataFrame(DFTable)
    DataframeTable[Prop].append(DF)


##########################BUILDING_EXCEL_SHEET#########################################
Sheet['A2'].value = "Name"
Sheet['A3'].value = "Prediction"
Sheet['A4'].value = "Line"
Sheet['A5'].value = "Difference"
Sheet['A6'].value = "O/U"
Sheet['A8'].value = "DK O"
Sheet['A9'].value = "DK U"
Sheet['A11'].value = "Last 10 Games"

for cols in Sheet.iter_cols(min_col=1, max_col=1, min_row=None):
    for cell in cols:
        cell.font = Font(bold=True)

RedFill = PatternFill(start_color='00F46666', end_color='00F46666', fill_type='solid')
GreenFill = PatternFill(start_color='0080FB7D', end_color='0080FB7D', fill_type='solid')
LightBlueFill = PatternFill(start_color='0081DEFF', end_color='0081DEFF', fill_type='solid')
GoldFill = PatternFill(start_color='00FFF061', end_color='00FFF061', fill_type='solid')
GrayFill = PatternFill(start_color='007F7F7F', end_color='007F7F7F', fill_type='solid')

for i in range(len(Categories)):
    prop = Categories[i]
    Header_cell = Sheet.cell(row=1, column=((3*i)+2))
    Header_cell.value = prop
    Header_cell.font = Font(bold=True)
    Sheet.merge_cells(start_row=1, start_column=((3*i)+2), end_row=1, end_column=((3*i)+4))

    DKNameElements = []
    if prop in DKLinks.keys():
        driver.get(DKLinks[prop])
        content = driver.page_source
        soup = BeautifulSoup(content, features="html.parser")

        try:
            wrapper = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'sportsbook-row-name')))
        except TimeoutException:
            print("Timeout")

        DKNameElements = driver.find_elements(By.CLASS_NAME, "sportsbook-row-name")

    Iters = len((Datatable[prop]))/2
    for j in range(int(Iters)):
        NameCell = Sheet.cell(row=2, column=((3*i) + 2 + j))
        PredictionCell = Sheet.cell(row=3, column=((3 * i) + 2 + j))
        LineCell = Sheet.cell(row=4, column=((3 * i) + 2 + j))
        DiffCell = Sheet.cell(row=5, column=((3 * i) + 2 + j))
        OUCell = Sheet.cell(row=6, column=((3 * i) + 2 + j))
        DKOCell = Sheet.cell(row=8, column=((3 * i) + 2 + j))
        DKUCell = Sheet.cell(row=9, column=((3 * i) + 2 + j))

        LineValue = float(Datatable[prop][2*j][1])
        PredictionValue = float(Datatable[prop][2*j][2])

        NameCell.value = Datatable[prop][2*j][0]
        LineCell.value = LineValue
        PredictionCell.value = PredictionValue
        DiffCell.value = abs(LineValue-PredictionValue)

        if PredictionValue > LineValue:
            OUCell.value = "O"
        else:
            OUCell.value = "U"

        if prop in DKLinks.keys():
            NameString = Datatable[prop][2*j][0]
            EditedNameString = NameString[:(NameString.find("(")-1)]

            for NameElement in DKNameElements:
                Name = NameElement.get_attribute("innerText")

                if Name == EditedNameString:
                    TRElement = NameElement.find_element(By.XPATH, ".//ancestor::tr")
                    OUElements = TRElement.find_elements(By.CLASS_NAME, "sportsbook-outcome-cell__element")
                    OValue = OUElements[1].get_attribute("innerText")
                    UValue = OUElements[3].get_attribute("innerText")

                    OMag = float(OValue[-3:])
                    OSign = 0
                    UMag = float(UValue[-3:])
                    USign = 0

                    if OValue[0] == "+":
                        OSign = 1
                    else:
                        OSign = -1

                    if UValue[0] == "+":
                        USign = 1
                    else:
                        USign = -1

                    OValue = OMag * OSign
                    UValue = UMag * USign

                    DKOCell.value = OValue
                    DKUCell.value = UValue

                    if float(OValue) < -130:
                        DKOCell.fill = LightBlueFill
                        if PredictionValue > LineValue:
                            NameCell.fill = GoldFill
                    if float(UValue) < -130:
                        DKUCell.fill = LightBlueFill
                        if PredictionValue < LineValue:
                            NameCell.fill = GoldFill

        for p in range(len(Datatable[prop][(2*j)+1])):
            c = Sheet.cell(row=(11+p), column=((3 * i) + 2 + j))
            HistValue = float(Datatable[prop][(2*j)+1][p])
            c.value = HistValue

            if PredictionValue > LineValue:
                if HistValue > LineValue:
                    c.fill = GreenFill
                elif HistValue < LineValue:
                    c.fill = RedFill
                else:
                    c.fill = GrayFill
            else:
                if HistValue < LineValue:
                    c.fill = GreenFill
                elif HistValue > LineValue:
                    c.fill = RedFill
                else:
                    c.fill = GrayFill


Workbook.save("C:\\Users\\harsh\\Downloads\\Code_Bet.xlsx")
